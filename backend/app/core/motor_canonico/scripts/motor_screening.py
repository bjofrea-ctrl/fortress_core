# -*- coding: utf-8 -*-
"""
Motor de Screening — Aprende a Invertir · Motor v1 r12
=====================================================
Filtro cuantitativo de acciones sobre un export del screener de InvestingPro.
Tres tribunales: CALIDAD (puerta) → SALUD FINANCIERA (cinturón) → PRECIO (decide).

REGLA DE ORO: este script es la ÚNICA fuente de clasificación. La lógica de
evaluate() es el motor canónico validado (réplica 1000/1000 en QA) y NO se
modifica sin re-validar contra la referencia.

Uso:  python motor_screening.py <export.xlsx> [--outdir DIR]
Salidas: informe ejecutivo por stdout · Screening_AAI_<fecha>.xlsx · Dashboard_AAI_<fecha>.html
"""
import sys, os, json, base64, argparse, html as _html
from datetime import date
from pathlib import Path
from collections import Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

MOTOR_VERSION = "Motor v1 r13"
ASSETS = Path(__file__).resolve().parent.parent / "assets"

# ───────────────────────── 1 · LECTURA POR NOMBRE (regla dura) ─────────────────────────
# El orden de columnas es libre: se localiza la fila de encabezados (contiene
# "Name" y "Ticker", esté donde esté), se construye un mapa nombre→columna
# normalizando espacios/saltos de línea, y TODO se lee por nombre exacto.

NA = {'-', 'NA', 'NM', 'N/A', 'UNAVAILABLE', 'NONE', '#RESTRICTED!', ''}

COLS_ESENCIALES = ['Name', 'Ticker', 'Price, Current']
COLS_NUCLEO = [
    'Fair Value', 'Fair Value Label (Analyst Targets)', 'Overall Health Label',
    'EV / EBIT', 'Free Cash Flow Yield',
    'Return on Invested Capital', 'Avg Return on Invested Capital (5y)',
    'Return on Equity', 'Gross Profit Margin',
    'Avg EPS Growth (5y)', 'Revenue CAGR (5y)',
    'FCF / Net Income', 'Buyback Yield',
    'Piotroski Score', 'Altman Z-Score', 'Beneish M-Score',
    'Total Debt / Total Capital',
]
COLS_OPCIONALES = ['Full Ticker', 'Market Cap (Adjusted)', 'P/E Ratio', 'PEG Ratio Fwd', 'Beta (5 Year)']

def _h(v):
    """Escapa texto proveniente del export antes de inyectarlo en el HTML del dashboard.
    Sin esto, nombres con & o < (p. ej. 'Jack Henry & Associates') generan HTML invalido."""
    return _html.escape(str(v), quote=True)

def _norm(v):
    return ' '.join(str(v).split()) if v is not None else ''

def leer_export(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    M = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    hr = None
    for i, row in enumerate(M):
        vals = {_norm(x) for x in row}
        if 'Name' in vals and 'Ticker' in vals:
            hr = i; break
    if hr is None:
        raise SystemExit("ERROR: no encontré la fila de encabezados (debe contener 'Name' y 'Ticker'). "
                         "Verifica que el archivo sea el export del screener de InvestingPro sin renombrar columnas.")
    hmap = {_norm(v): j for j, v in enumerate(M[hr]) if _norm(v)}
    faltan_esc = [c for c in COLS_ESENCIALES if c not in hmap]
    if faltan_esc:
        raise SystemExit(f"ERROR: faltan columnas esenciales: {faltan_esc}. No se puede correr el filtro.")
    faltantes = [c for c in COLS_NUCLEO if c not in hmap]
    filas = [row for row in M[hr + 1:] if row and _norm(row[hmap['Name']]) != '']
    return filas, hmap, faltantes

# ───────────────────────── 2 · MOTOR CANÓNICO (verbatim del QA validado) ─────────────────────────

FUND_KW = ['ETF', 'FUND', 'TRUST', 'ISHARES', 'SPDR', 'VANGUARD', 'INVESCO', 'INDEX']

# Orden de severidad de las alertas: la columna muestra las 3 primeras, así que
# las de riesgo real (quiebra, contabilidad) nunca pueden quedar tapadas por las informativas.
EMOJI_BALDE = {'Deep Dive':'🔬 Deep Dive','Watchlist':'📋 Watchlist','Neutral':'⚪ Neutral','Descartada':'❌ Descartada'}
PREF_ALERTA = {'Posible trampa de valor':'⚠️ ','¿Demasiado barata?':'⚠️ ','M-Score':'⚠️ ','Altman zona de riesgo':'⚠️ ',
               'Excelente pero salud débil':'⚠️ ','Datos insuficientes':'⚠️ ','Revisar ROIC/moat':'🔎 ','Mejora débil':'🔎 '}

ORDEN_ALERTAS = ['Datos insuficientes', 'Altman zona de riesgo', 'M-Score', 'Posible trampa de valor',
                 'Excelente pero salud débil', '¿Demasiado barata?', 'Mejora débil']

def _cargar_sectores_excluidos():
    """Lista de tickers de Financieras/Utilities (referencia InvestingPro) que el método no mide.
    La lista blanca de plataformas de pago ya viene aplicada en el CSV.
    Si el archivo falta o viene vacio, el motor ABORTA: sin esta lista la exclusion sectorial
    no se aplica y la clasificacion sale distinta sin previo aviso."""
    p = Path(__file__).resolve().parent.parent / 'references' / 'sectores_excluidos.csv'
    if not p.exists():
        raise SystemExit(
            f"ERROR CRITICO: no se encontro la lista de exclusion sectorial en {p}\n"
            "El motor debe ejecutarse desde la estructura completa del skill "
            "(scripts/ junto a references/ y assets/). Sin ese archivo las empresas "
            "Financieras y de Utilities NO se excluyen y la clasificacion resultante es invalida."
        )
    s = {}
    try:
        with open(p, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#') or line.startswith('ticker,'):
                    continue
                parts = line.split(',')
                if len(parts) >= 3:
                    s[parts[0].strip()] = parts[-1].strip()
    except OSError as e:
        raise SystemExit(f"ERROR CRITICO: no se pudo leer {p} ({e}). Clasificacion abortada.")
    if not s:
        raise SystemExit(
            f"ERROR CRITICO: la lista de exclusion sectorial en {p} esta vacia o mal formada. "
            "Clasificacion abortada para no entregar resultados invalidos."
        )
    return s

SECT_EXCL = _cargar_sectores_excluidos()

COLS_PCT = ['Return on Invested Capital', 'Avg Return on Invested Capital (5y)', 'Return on Equity',
            'Gross Profit Margin', 'Free Cash Flow Yield', 'Avg EPS Growth (5y)', 'Revenue CAGR (5y)',
            'Buyback Yield', 'Total Debt / Total Capital']

def verificar_convencion(filas, hmap):
    """Guardia ruidosa: si InvestingPro cambiara el formato de sus porcentajes (decimal → %),
    todos los umbrales quedarían 100x descalibrados. Devuelve las columnas sospechosas."""
    sosp = []
    for c in COLS_PCT:
        j = hmap.get(c)
        if j is None: continue
        vs = sorted(abs(v) for v in (row[j] if j < len(row) else None for row in filas)
                    if isinstance(v, (int, float)))
        if len(vs) >= 20 and vs[len(vs) // 2] >= 5:
            sosp.append(c)
    return sosp

def _mk_accessors(hmap):
    def raw(row, h):
        j = hmap.get(h)
        if j is None or j >= len(row): return None
        v = row[j]
        if v is None: return None
        if isinstance(v, str) and v.strip().upper() in NA: return None
        return v
    def num(row, h):
        v = raw(row, h); return float(v) if isinstance(v, (int, float)) else None
    def pctv(row, h):
        # El export de InvestingPro entrega los porcentajes SIEMPRE como decimales
        # (0.478 = 47.8%). Verificado columna por columna en exports independientes.
        # No se adivina por celda: adivinar rompía los ROE > 500% (patrimonio pequeño
        # por recompras), leyéndolos como porcentajes diminutos. Ver verificar_convencion().
        v = num(row, h)
        return None if v is None else v * 100
    return raw, num, pctv

def evaluate(row, raw, num, pctv):
    nm = str(raw(row, 'Name') or '')
    pio = num(row, 'Piotroski Score'); roic = pctv(row, 'Return on Invested Capital')
    roe = pctv(row, 'Return on Equity'); alt = num(row, 'Altman Z-Score')
    if (any(k in nm.upper() for k in FUND_KW) and (pio is None or roic is None)) or all(x is None for x in [roic, roe, pio, alt]):
        return dict(fund=True, name=nm, ticker=str(raw(row, 'Ticker') or ''))
    _tk0 = str(raw(row, 'Ticker') or '').strip()
    if _tk0 and _tk0 in SECT_EXCL:
        return dict(fund=True, sectx=SECT_EXCL[_tk0], name=nm, ticker=_tk0)
    roic5 = pctv(row, 'Avg Return on Invested Capital (5y)'); gm = pctv(row, 'Gross Profit Margin')
    fcfy = pctv(row, 'Free Cash Flow Yield'); epsg = pctv(row, 'Avg EPS Growth (5y)'); revg = pctv(row, 'Revenue CAGR (5y)')
    bb = pctv(row, 'Buyback Yield'); dc = pctv(row, 'Total Debt / Total Capital'); fcfni = num(row, 'FCF / Net Income')
    evebit = num(row, 'EV / EBIT'); fv = num(row, 'Fair Value'); price = num(row, 'Price, Current'); m = num(row, 'Beneish M-Score')
    health = str(raw(row, 'Overall Health Label') or '').strip().lower()
    lbl = str(raw(row, 'Fair Value Label (Analyst Targets)') or '').strip().lower()
    upside = (fv - price) / price if (fv is not None and price and price > 0) else None
    # Lente 1 — Greenblatt (ROIC actual + 5y)
    if roic is None and roic5 is None: L1 = 'N/D'
    elif roic is not None and roic5 is not None:
        L1 = 'Sí' if (roic >= 20 and roic5 >= 20) else ('No' if (roic < 10 or roic5 < 10) else 'Medio')
    else:
        v = roic if roic is not None else roic5
        L1 = 'Medio' if v >= 20 else ('No' if v < 10 else 'Medio')
    # Lente 2 — MSCI (ROE + apalancamiento sano)
    lev = (dc is not None and dc > 60) or (dc is None and alt is not None and alt < 1.81)
    if roe is None: L3 = 'N/D'
    elif roe >= 15 and not lev: L3 = 'Sí'
    elif roe < 10 or lev: L3 = 'No'
    else: L3 = 'Medio'
    # Lente 3 — AQR (4 pilares; RENTABLE nunca marca ✗)
    conv = (fcfni is not None and fcfni >= 0.8 and fcfy is not None and fcfy > 0)
    rent = 'v' if sum([roe is not None and roe >= 15, gm is not None and gm >= 40, conv]) >= 2 else 'i'
    if epsg is None and revg is None: cre = 'i'
    elif epsg is not None and revg is not None: cre = 'v' if (epsg > 0 and revg > 0) else ('x' if (epsg < 0 or revg < 0) else 'i')
    else:
        g = epsg if epsg is not None else revg; cre = 'v' if g > 0 else ('x' if g < 0 else 'i')
    if alt is None: seg = 'i'
    elif alt >= 3: seg = 'v'
    elif alt < 1.81: seg = 'x'
    else: seg = 'v' if health in ('good', 'great', 'excellent') else 'i'
    if bb is None: pay = 'i'
    elif bb >= 0: pay = 'v'
    elif bb < -2: pay = 'x'
    else: pay = 'i'
    pil = [rent, cre, seg, pay]
    L4 = 'Sí' if (pil.count('v') >= 3 and pil.count('x') == 0) else ('No' if pil.count('x') >= 2 else 'Medio')
    lentes = [L1, L3, L4]; n_cal = lentes.count('Sí')
    cal = 'EXCELENTE' if n_cal == 3 else ('BUENA' if n_cal == 2 else 'DÉBIL')
    # PRECIO — veredicto neto
    n_val = sum([evebit is not None and 0 < evebit <= 10, upside is not None and upside >= 0.20,
                 fcfy is not None and fcfy >= 4, lbl in ('bargain', 'undervalued')])
    v_c = sum([upside is not None and upside <= -0.10, evebit is not None and (evebit >= 20 or evebit < 0),
               fcfy is not None and fcfy <= 2, lbl == 'overvalued'])
    if v_c >= 2 and v_c > n_val: precio = 'CARA'
    elif n_val >= 3 and n_val > v_c: precio = 'MUY BARATA'
    elif n_val == 2 and n_val > v_c: precio = 'BARATA'
    else: precio = 'MIXTA'
    # SALUD
    n_sal = sum([pio is not None and pio >= 7, alt is not None and alt >= 3, m is not None and m <= -1.78])
    salud = 'DÉBIL' if ((alt is not None and alt < 1.81) or n_sal <= 1) else ('EXCELENTE' if n_sal == 3 else 'MIXTA')
    # Cobertura + categorías
    core = [roic, roic5, roe, gm, epsg, revg, fcfni, bb, evebit, fcfy, fv, pio]
    insuf = sum(1 for x in core if x is None) >= 5
    trampa = (cal == 'DÉBIL' and precio in ('MUY BARATA', 'BARATA'))
    if insuf: balde = 'Descartada'
    elif cal == 'EXCELENTE' and salud != 'DÉBIL' and precio in ('MUY BARATA', 'BARATA'): balde = 'Deep Dive'
    elif cal == 'EXCELENTE' and salud != 'DÉBIL': balde = 'Watchlist'
    elif cal in ('EXCELENTE', 'BUENA'): balde = 'Neutral'
    else: balde = 'Descartada'
    al = []
    if trampa: al.append('Posible trampa de valor')
    if cal == 'EXCELENTE' and ((upside is not None and upside >= 0.40) or n_val >= 3): al.append('¿Demasiado barata?')
    if m is not None and m > -1.78: al.append('M-Score')
    if cal == 'EXCELENTE' and pio is not None and pio <= 3: al.append('Mejora débil')
    if alt is not None and alt < 1.81: al.append('Altman zona de riesgo')
    if cal == 'EXCELENTE' and salud == 'DÉBIL': al.append('Excelente pero salud débil')
    if insuf: al.append('Datos insuficientes')
    al.sort(key=lambda x: ORDEN_ALERTAS.index(x) if x in ORDEN_ALERTAS else 99)
    reds = set(); amb = set()
    if roic is not None and roic < 10: reds.add('Return on Invested Capital')
    if roic5 is not None and roic5 < 10: reds.add('Avg Return on Invested Capital (5y)')
    if roe is not None and roe < 10: reds.add('Return on Equity')
    if epsg is not None and epsg < 0: reds.add('Avg EPS Growth (5y)')
    if revg is not None and revg < 0: reds.add('Revenue CAGR (5y)')
    if fcfni is not None and not conv and (fcfni < 0.8 or (fcfy is not None and fcfy <= 0)): reds.add('FCF / Net Income')
    if bb is not None and bb < -2: reds.add('Buyback Yield')
    if dc is not None and dc > 60: reds.add('Total Debt / Total Capital')
    if pio is not None and pio <= 3: reds.add('Piotroski Score')
    if alt is not None:
        if alt < 1.81: reds.add('Altman Z-Score')
        elif alt < 3: amb.add('Altman Z-Score')
    if m is not None and m > -1.78: amb.add('Beneish M-Score')
    if fcfy is not None and fcfy <= 2: reds.add('Free Cash Flow Yield')
    if evebit is not None and (evebit >= 20 or evebit < 0): reds.add('EV / EBIT')
    return dict(fund=False, name=nm, ticker=str(raw(row, 'Ticker') or ''),
                balde=balde, punt=n_cal + n_val + n_sal, L=lentes, cal=cal, n_cal=n_cal,
                precio=precio, n_val=n_val, v_c=v_c, upside=upside, salud=salud, n_sal=n_sal,
                alertas=' · '.join(al[:3]) if al else '—', reds=reds, amb=amb,
                up_red=(upside is not None and upside <= -0.10),
                mets=dict(mcap=num(row, 'Market Cap (Adjusted)'),
                          roic=roic, roic5=roic5, roe=roe, gm=gm, fcfy=fcfy, evebit=evebit,
                          pio=pio, alt=alt, m=m, dc=dc, fv=fv, price=price, epsg=epsg,
                          revg=revg, fcfni=fcfni, bb=bb, lbl=lbl, ft=str(raw(row, 'Full Ticker') or '')))

# ───────────────────────── 3 · EXCEL ENRIQUECIDO (spec v5 + marca) ─────────────────────────

def generar_excel(filas, hmap, evals, orden, outpath, export_name, fecha):
    raw, num, pctv = _mk_accessors(hmap)
    def col(h, g, a): return (h, g, a)
    layout = [col('Balde','RESULTADO',True), col('Puntaje (0-10)','RESULTADO',True),
        col('Name','EMPRESA',False), col('Ticker','EMPRESA',False), col('Price, Current','EMPRESA',False), col('Market Cap (Adjusted)','EMPRESA',False),
        col('Return on Invested Capital','CALIDAD',False), col('Avg Return on Invested Capital (5y)','CALIDAD',False), col('Return on Equity','CALIDAD',False),
        col('Gross Profit Margin','CALIDAD',False), col('Avg EPS Growth (5y)','CALIDAD',False), col('Revenue CAGR (5y)','CALIDAD',False),
        col('FCF / Net Income','CALIDAD',False), col('Buyback Yield','CALIDAD',False), col('Total Debt / Total Capital','CALIDAD',False), col('Beta (5 Year)','CALIDAD',False),
        col('Greenblatt','CALIDAD',True), col('MSCI','CALIDAD',True), col('AQR','CALIDAD',True), col('Veredicto Calidad','CALIDAD',True), col('Señales de calidad (0-3)','CALIDAD',True),
        col('P/E Ratio','PRECIO',False), col('PEG Ratio Fwd','PRECIO',False), col('EV / EBIT','PRECIO',False), col('Free Cash Flow Yield','PRECIO',False),
        col('Fair Value','PRECIO',False), col('Fair Value Label (Analyst Targets)','PRECIO',False),
        col('Upside vs Fair Value','PRECIO',True), col('Veredicto Precio','PRECIO',True), col('Señales de precio (0-4)','PRECIO',True),
        col('Piotroski Score','SALUD',False), col('Altman Z-Score','SALUD',False), col('Beneish M-Score','SALUD',False), col('Overall Health Label','SALUD',False),
        col('Veredicto Salud financiera','SALUD',True), col('Señales de salud (0-3)','SALUD',True),
        col('Alertas','RESULTADO',True)]
    BAND={'RESULTADO':'404040','EMPRESA':'595959','CALIDAD':'375623','PRECIO':'7F6000','SALUD':'1F4E79'}
    DARK={'RESULTADO':'404040','EMPRESA':'595959','CALIDAD':'538135','PRECIO':'BF8F00','SALUD':'2E75B6'}
    LIGHT={'RESULTADO':'D9D9D9','EMPRESA':'D9D9D9','CALIDAD':'C6E0B4','PRECIO':'FFE699','SALUD':'BDD7EE'}
    BANDLBL={'RESULTADO':'RESULTADO','EMPRESA':'EMPRESA','CALIDAD':'CALIDAD — ¿es un gran negocio?','PRECIO':'PRECIO — ¿está barata?','SALUD':'SALUD FINANCIERA — ¿sólida y limpia?'}
    PCT={'Free Cash Flow Yield','Return on Invested Capital','Avg Return on Invested Capital (5y)','Return on Equity','Gross Profit Margin','Avg EPS Growth (5y)','Revenue CAGR (5y)','FCF / Net Income','Buyback Yield','Total Debt / Total Capital','Upside vs Fair Value'}
    NUM2={'Price, Current','Fair Value','P/E Ratio','PEG Ratio Fwd','EV / EBIT','Altman Z-Score','Beneish M-Score','Beta (5 Year)'}
    # Columnas del export (no calculadas) que no llegaron: se dibujan igual, vacías y en rojo.
    FALTAN=[h for (h,g,a) in layout if not a and h not in hmap]
    OFF=1 if FALTAN else 0          # desplaza bandas/encabezados/datos para dejar el letrero arriba
    wb=Workbook(); s=wb.active; s.title='Screening'
    BASE=Font(name='Arial',size=10)
    RED=PatternFill('solid',start_color='FFC7CE'); REDF=Font(name='Arial',size=10,color='9C0006')
    AMB=PatternFill('solid',start_color='FFEB9C'); AMBF=Font(name='Arial',size=10,color='9C6500')
    GRN=PatternFill('solid',start_color='C6EFCE'); GRNF=Font(name='Arial',size=10,color='006100')
    GRN2=PatternFill('solid',start_color='A9D08E'); GRN2F=Font(name='Arial',size=10,color='004B00',bold=True)
    bfill={'Deep Dive':PatternFill('solid',start_color='A9D08E'),'Watchlist':PatternFill('solid',start_color='E2EFDA'),'Neutral':PatternFill('solid',start_color='FFF2CC'),'Descartada':PatternFill('solid',start_color='FFC7CE'),'Fondo/ETF (excluido)':PatternFill('solid',start_color='E0E0E0'),'Omitida':PatternFill('solid',start_color='D6DEE8')}
    i=0
    while i<len(layout):
        g=layout[i][1]; j=i
        while j<len(layout) and layout[j][1]==g: j+=1
        s.merge_cells(f"{get_column_letter(i+1)}{1+OFF}:{get_column_letter(j)}{1+OFF}")
        cell=s.cell(row=1+OFF,column=i+1,value=BANDLBL[g])
        cell.fill=PatternFill('solid',start_color=BAND[g]); cell.font=Font(name='Arial',bold=True,color='FFFFFF',size=10); cell.alignment=Alignment(horizontal='center',vertical='center')
        i=j
    for j,(h,g,a) in enumerate(layout,1):
        c=s.cell(row=2+OFF,column=j,value=h.replace('Market Cap (Adjusted)','Market Cap (US$ B)').replace('Upside vs Fair Value','Price vs Fair Value'))
        if a: c.fill=PatternFill('solid',start_color=DARK[g]); c.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
        else: c.fill=PatternFill('solid',start_color=LIGHT[g]); c.font=Font(name='Arial',bold=True,color='262626',size=10)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    names={h:j for j,(h,g,a) in enumerate(layout,1)}
    s.cell(row=2+OFF,column=names['Puntaje (0-10)']).comment=Comment('Puntaje = señales de calidad (0-3) + señales de precio a favor (0-4) + señales de salud (0-3). Máximo 10.','Método')
    s.cell(row=2+OFF,column=names['Veredicto Calidad']).comment=Comment('3 lentes: Greenblatt (ROIC), MSCI (ROE+deuda), AQR (4 pilares). EXCELENTE = 3 lentes en Sí · BUENA = 2 · DÉBIL = 0-1. La calidad es la puerta: sin EXCELENTE no hay Deep Dive ni Watchlist.','Método')
    s.cell(row=2+OFF,column=names['Veredicto Precio']).comment=Comment('4 jueces: EV/EBIT ≤10 (y positivo) · descuento ≥20% vs Fair Value · FCF Yield ≥4% · etiqueta de analistas. MUY BARATA = 3-4 a favor · BARATA = 2 · MIXTA = 1 o empate · CARA = más en contra que a favor. Veredicto neto, siempre.','Método')
    s.cell(row=2+OFF,column=names['Veredicto Salud financiera']).comment=Comment('3 señales: Piotroski ≥7 · Altman ≥3 · M-Score sin alerta. EXCELENTE = 3 · MIXTA = 2 · DÉBIL = 0-1 o Altman <1.81. Salud DÉBIL bloquea Deep Dive y Watchlist.','Método')
    for _hn,_tx in [('Piotroski Score','Solidez financiera de 0 a 9. ≥7 suma una señal de salud · ≤3 celda roja (y si la calidad es excelente, alerta de mejora débil).'),
                    ('Altman Z-Score','Riesgo de quiebra. ≥3 suma señal de salud · entre 1.81 y 3 zona gris (ámbar) · <1.81 zona de riesgo: la salud queda DÉBIL automáticamente.'),
                    ('Beneish M-Score','Detector de manipulación contable. ≤ −1.78 limpio y suma señal · > −1.78 alerta (ámbar): patrón compatible con resultados maquillados. No es una prueba, es un motivo de revisión.'),
                    ('Return on Invested Capital','Lente Greenblatt: ≥20% (actual y promedio de 5 años) = lente en Sí · <10% = celda roja.'),
                    ('Avg Return on Invested Capital (5y)','Se compara con el ROIC actual: si el actual está por debajo del promedio, salta la alerta Revisar ROIC/moat (posible erosión de la ventaja competitiva).'),
                    ('Return on Equity','Lente MSCI: ≥15% con deuda sana = lente en Sí · <10% = celda roja.'),
                    ('Gross Profit Margin','Pilar RENTABLE del lente AQR: ≥40% cuenta a favor.'),
                    ('EV / EBIT','Juez de precio: ≤10 y positivo cuenta a favor · ≥20 o negativo cuenta en contra (celda roja).'),
                    ('Free Cash Flow Yield','Juez de precio: ≥4% a favor · ≤2% en contra (celda roja).'),
                    ('Fair Value','Valor justo de InvestingPro. Juez de precio: un descuento ≥20% frente al precio actual cuenta a favor.'),
                    ('Total Debt / Total Capital','>60% = apalancada (celda roja) y tumba el lente MSCI. En bancos este número no significa lo mismo: por eso el método los omite.')]:
        if _hn in names: s.cell(row=2+OFF,column=names[_hn]).comment=Comment(_tx,'Método')
    ri=3+OFF
    for k in orden:
        e=evals[k]; row=filas[k]
        mo = (not e['fund']) and e['mets']['roic'] is not None and e['mets']['roic5'] is not None and e['mets']['roic'] < e['mets']['roic5']
        for j,(h,g,addf) in enumerate(layout,1):
            if e['fund']:
                if h=='Balde': v=f"⚙️ Omitida por método ({e['sectx']})" if e.get('sectx') else 'Fondo/ETF (excluido)'
                elif h=='Alertas': v=f"Sector {e['sectx']} — este método no lo mide bien; analizar aparte" if e.get('sectx') else 'Fondo/ETF — fuera del análisis'
                elif not addf:
                    v=raw(row,h)
                    if h=='Market Cap (Adjusted)' and isinstance(v,(int,float)): v=v/1e9
                else: v=None
            else:
                if h=='Balde': v=EMOJI_BALDE.get(e['balde'],e['balde'])
                elif h=='Puntaje (0-10)': v=e['punt']
                elif h=='Greenblatt': v=e['L'][0]
                elif h=='MSCI': v=e['L'][1]
                elif h=='AQR': v=e['L'][2]
                elif h=='Veredicto Calidad': v=e['cal']
                elif h=='Señales de calidad (0-3)': v=e['n_cal']
                elif h=='Upside vs Fair Value': v=e['upside']
                elif h=='Veredicto Precio': v=e['precio']
                elif h=='Señales de precio (0-4)': v=e['n_val']
                elif h=='Veredicto Salud financiera': v=e['salud']
                elif h=='Señales de salud (0-3)': v=e['n_sal']
                elif h=='Alertas':
                    v=e['alertas']
                    if mo and e['balde'] in ('Deep Dive','Watchlist'):
                        v='Revisar ROIC/moat' if v=='—' else v+' · Revisar ROIC/moat'
                    if v!='—':
                        v=' · '.join(PREF_ALERTA.get(tk,'')+tk for tk in v.split(' · '))
                else:
                    v=raw(row,h)
                    if h=='Market Cap (Adjusted)' and isinstance(v,(int,float)): v=v/1e9
            c=s.cell(row=ri,column=j,value=v)
            fill=None; font=BASE
            if not e['fund']:
                if h in e['reds'] or (h=='Upside vs Fair Value' and e['up_red']): fill,font=RED,REDF
                elif h in e['amb']: fill,font=AMB,AMBF
                elif h=='Return on Invested Capital' and mo: fill,font=AMB,AMBF
                elif h in ('Greenblatt','MSCI','AQR'):
                    if v=='Sí': fill,font=GRN,GRNF
                    elif v=='Medio': fill,font=AMB,AMBF
                    elif v=='No': fill,font=RED,REDF
                elif h=='Veredicto Calidad':
                    if v=='EXCELENTE': fill,font=GRN,GRNF
                    elif v=='BUENA': fill,font=AMB,AMBF
                    elif v=='DÉBIL': fill,font=RED,REDF
                elif h=='Veredicto Precio':
                    if v=='MUY BARATA': fill,font=GRN2,GRN2F
                    elif v=='BARATA': fill,font=GRN,GRNF
                    elif v=='CARA': fill,font=RED,REDF
                elif h=='Veredicto Salud financiera':
                    if v=='EXCELENTE': fill,font=GRN,GRNF
                    elif v=='MIXTA': fill,font=AMB,AMBF
                    elif v=='DÉBIL': fill,font=RED,REDF
                elif h=='Alertas' and v!='—': fill,font=AMB,AMBF
            c.font=font
            if fill is not None: c.fill=fill
            if h in PCT and isinstance(v,(int,float)): c.number_format='0.0%'
            if h=='Market Cap (Adjusted)' and isinstance(v,(int,float)): c.number_format='#,##0.0'
            if h in NUM2 and isinstance(v,(int,float)): c.number_format='0.00'
            if h in ('Greenblatt','MSCI','AQR','Señales de calidad (0-3)','Señales de precio (0-4)','Señales de salud (0-3)','Puntaje (0-10)'): c.alignment=Alignment(horizontal='center')
        s.cell(row=ri,column=1).fill=bfill[('Omitida' if e.get('sectx') else 'Fondo/ETF (excluido)') if e['fund'] else e['balde']]
        ri+=1
    s.freeze_panes=f'E{3+OFF}'
    s.auto_filter.ref=f"A{2+OFF}:{get_column_letter(len(layout))}{ri-1}"
    Lup=get_column_letter(names['Upside vs Fair Value'])
    rng_barras=f"{Lup}{3+OFF}:{Lup}{ri-1}"
    s.conditional_formatting.add(rng_barras, DataBarRule(start_type='min', end_type='max', color='63C384', showValue=True))
    if FALTAN:
        # letrero rojo a todo lo ancho
        s.merge_cells(f"A1:{get_column_letter(len(layout))}1")
        _nuc=[h for h in FALTAN if h in COLS_NUCLEO]
        _cola=("Las columnas en rojo llegaron vacías y SÍ afectan el resultado: varias empresas quedaron clasificadas por debajo de lo que les corresponde. Esta corrida NO es comparable con una completa."
               if _nuc else "Las columnas en rojo llegaron vacías. No cambian la clasificación, pero pierdes contexto en el análisis.")
        _b=s.cell(row=1,column=1,value=f"⚠️  FALTAN {len(FALTAN)} INDICADORES EN TU EXPORT — agrégalos en el screener de InvestingPro y vuelve a exportar:  {' · '.join(FALTAN)}    |    {_cola}")
        _b.fill=PatternFill('solid',start_color='C00000')
        _b.font=Font(name='Arial',bold=True,color='FFFFFF',size=11)
        _b.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
        s.row_dimensions[1].height=32
        # columna completa en rojo, con su encabezado marcado
        _FR=PatternFill('solid',start_color='FFC7CE'); _FF=Font(name='Arial',size=10,color='9C0006',italic=True)
        for _h in FALTAN:
            _j=names[_h]
            _hc=s.cell(row=2+OFF,column=_j)
            _hc.value=f"{_h}  ⚠️"
            _hc.fill=PatternFill('solid',start_color='C00000')
            _hc.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
            _hc.comment=Comment(f'Esta columna NO venía en tu export. Agrégala en el screener de InvestingPro con el nombre exacto "{_h}" y vuelve a exportar: sin ella, las empresas pierden señales y varias quedan clasificadas por debajo de lo que les corresponde.','Método')
            for _r in range(3+OFF,ri):
                _c=s.cell(row=_r,column=_j); _c.value='falta'; _c.fill=_FR; _c.font=_FF
                _c.alignment=Alignment(horizontal='center',vertical='center')
    s.column_dimensions['A'].width=16; s.column_dimensions['B'].width=9; s.column_dimensions['C'].width=30
    for j in range(4,len(layout)+1): s.column_dimensions[get_column_letter(j)].width=11
    _ca=get_column_letter(len(layout)); s.column_dimensions[_ca].width=46
    for _r in range(3+OFF,ri): s[f'{_ca}{_r}'].alignment=Alignment(wrap_text=True,vertical='center')
    s.row_dimensions[1].height=18; s.row_dimensions[2].height=30

    # ── Instructivo (primera hoja) con logo de marca ──
    noF=[k for k in evals if not evals[k]['fund']]
    cts=Counter(evals[k]['balde'] for k in noF)
    funds=sum(1 for k in evals if evals[k]['fund'] and not evals[k].get('sectx'))
    sectx=sum(1 for k in evals if evals[k].get('sectx'))
    exc=sum(1 for k in noF if evals[k]['cal']=='EXCELENTE' and evals[k]['salud']!='DÉBIL')
    tr=sum(1 for k in noF if evals[k]['alertas'].find('Posible trampa')>=0)
    ins=wb.create_sheet('Instructivo')
    ins.sheet_view.showGridLines=False
    NV='0A1B2D'; PT='003850'; CD='0F2A3F'; AM='F59C00'; TE='6DADAA'; OR2='F1801B'; CO='E07068'; WH='FFFFFF'; MU='A9B7C2'; CL='D8E0E6'
    ins.column_dimensions['A'].width=2.5; ins.column_dimensions['B'].width=25
    ins.column_dimensions['C'].width=104; ins.column_dimensions['D'].width=2.5
    _nvf=PatternFill('solid',start_color=NV)
    for _r in range(1,190):
        for _c in range(1,5): ins.cell(row=_r,column=_c).fill=_nvf
    def _cl(rw,col,val,color=WH,bold=False,size=10,fill=None,ital=False,align='left'):
        c=ins.cell(row=rw,column=col,value=val)
        c.font=Font(name='Arial',color=color,bold=bold,size=size,italic=ital)
        if fill: c.fill=PatternFill('solid',start_color=fill)
        c.alignment=Alignment(wrap_text=True,vertical='center',horizontal=align)
        return c
    def _alto(rw,txt,base=15):
        ins.row_dimensions[rw].height=max(base,13*(len(str(txt))//96+1)+3)
    R=2
    try:
        from openpyxl.drawing.image import Image as XLImage
        _lg=XLImage(str(ASSETS/'logo_completo_blanco.png')); _rt=_lg.width/_lg.height
        _lg.height=40; _lg.width=int(40*_rt); ins.add_image(_lg,'C2')
        ins.row_dimensions[2].height=34; R=4
    except Exception: R=2
    _cl(R,3,'GUÍA DEL ARCHIVO',color=AM,bold=True,size=10); R+=1
    _cl(R,3,'Screening cuantitativo — Aprende a Invertir',bold=True,size=16); ins.row_dimensions[R].height=24; R+=1
    _cl(R,3,f'Export: {export_name} · Corrida: {fecha} · {MOTOR_VERSION}',color=MU,size=9,ital=True); R+=2
    def SEC(rw,num,tit,acc):
        _cl(rw,2,num,color=NV,bold=True,size=13,fill=acc,align='center')
        _cl(rw,3,'  '+tit,color=WH,bold=True,size=12,fill=PT)
        ins.row_dimensions[rw].height=22
        return rw+1
    def FILA(rw,lab,txt,labfill=CD,labcolor=WH):
        _cl(rw,2,lab,color=labcolor,bold=True,size=9,fill=labfill,align='center')
        _cl(rw,3,txt,color=CL,size=10,fill=CD); _alto(rw,txt)
        return rw+1
    def TXT(rw,txt):
        _cl(rw,3,txt,color=CL,size=10,fill=CD); _alto(rw,txt)
        return rw+1
    R=SEC(R,'01','QUÉ ES ESTA HERRAMIENTA',AM)
    R=TXT(R,'Un filtro cuantitativo que toma tu export del screener y clasifica cada empresa con tres tribunales — CALIDAD, SALUD FINANCIERA y PRECIO — usando principios documentados de marcos reconocidos (Greenblatt, Piotroski, Altman, Beneish, MSCI Quality, AQR y valuación por margen de seguridad). Su trabajo es reducir ~1.000 acciones a una lista corta que merezca tu tiempo.')
    R=TXT(R,f'En esta corrida: {len(noF)+sectx} empresas → {exc} con calidad excelente Y salud sólida → {cts["Deep Dive"]} en Deep Dive · {cts["Watchlist"]} en Watchlist · {cts["Neutral"]} en Neutral · {cts["Descartada"]} Descartadas ({tr} con nota de posible trampa) · {funds} fondos/ETFs' + (f' · {sectx} omitidas por método.' if sectx else '.'))
    R+=1
    R=SEC(R,'02','LA FILOSOFÍA — POR QUÉ ESTE ORDEN',AM)
    R=TXT(R,'"Es mucho mejor comprar una empresa maravillosa a un precio justo que una empresa justa a un precio maravilloso" — Warren Buffett, carta a accionistas de 1989. Este filtro aplica esa idea en orden: primero la CALIDAD como puerta eliminatoria (sin excelencia, no hay nada que mirar), la SALUD FINANCIERA como cinturón de seguridad (anti-quiebra y anti-manipulación contable), y al final el PRECIO decide la lista. Es la evolución del value investing de Buffett y Munger — calidad primero — no el enfoque clásico de comprar cualquier cosa barata.')
    R+=1
    R=SEC(R,'03','LAS CATEGORÍAS',TE)
    R=FILA(R,'🔬 DEEP DIVE','Excelente empresa a excelente precio — SEGÚN LOS NÚMEROS. El nombre es literal: investigar a fondo antes de cualquier decisión. Los números son la mitad de la tarea; el análisis cualitativo puede cambiar la tesis por completo.','A9D08E',NV)
    R=FILA(R,'📋 WATCHLIST','Excelente empresa, pero el precio aún no es atractivo. Va al radar: se monitorea esperando un mejor precio.','E2EFDA',NV)
    R=FILA(R,'⚪ NEUTRAL','Ni excelente ni mala según los datos. Hoy no es candidata; puede graduarse si su calidad mejora.','FFF2CC',NV)
    R=FILA(R,'❌ DESCARTADA','Ni la calidad ni el precio son atractivos según los datos de hoy. NO es un pronóstico: una descartada puede subir de precio. Solo significa que, bajo esta metodología y hoy, no es interesante.','FFC7CE',NV)
    R=FILA(R,'⚙️ OMITIDA','Financieras y utilities: el método no las mide bien (en los bancos la deuda es la materia prima y en las utilities el retorno está regulado). Se apartan para no mentirte — analízalas con otro criterio. Las plataformas de pago como Visa, Mastercard o PayPal SÍ se analizan.','D6DEE8',NV)
    R=FILA(R,'FONDO / ETF','No es una empresa: queda fuera del análisis.','E0E0E0',NV)
    R+=1
    R=SEC(R,'04','LAS ALERTAS — QUÉ SIGNIFICA CADA UNA',OR2)
    R=TXT(R,'La columna Alertas muestra hasta tres, ordenadas de mayor a menor severidad.')
    R=FILA(R,'⚠️ Datos insuficientes','Faltan cinco o más métricas núcleo: la empresa se descarta por falta de información, no por ser mala.')
    R=FILA(R,'⚠️ Altman zona de riesgo','Z-Score por debajo de 1.81: riesgo financiero elevado. La salud queda en DÉBIL automáticamente.')
    R=FILA(R,'⚠️ M-Score','Beneish por encima de −1.78: patrón compatible con manipulación contable. No es una prueba — es un motivo de revisión.')
    R=FILA(R,'⚠️ Posible trampa de valor','Empresa débil que se ve barata. Lo barato es el anzuelo: cuando una empresa floja está muy descontada, el descuento suele tener dueño. Doble cuidado.')
    R=FILA(R,'⚠️ Excelente pero salud débil','Pasó el tribunal de calidad pero reprobó el de salud: el cinturón de seguridad la frenó antes de llegar a Deep Dive o Watchlist.')
    R=FILA(R,'⚠️ ¿Demasiado barata?','Excelente empresa con un descuento inusualmente grande. La pregunta obligada: ¿qué sabe el mercado que los números no muestran?')
    R=FILA(R,'🔎 Revisar ROIC/moat','El ROIC actual está por debajo de su promedio de 5 años: posible erosión de la ventaja competitiva. Puede ser un mal año o una gran inversión en curso — investiga el porqué.')
    R=FILA(R,'🔎 Mejora débil','Excelente en calidad pero Piotroski ≤3: la foto es buena, la tendencia interna no acompaña.')
    R+=1
    R=SEC(R,'05','EL PUNTAJE Y LOS TRES VEREDICTOS',AM)
    R=FILA(R,'Puntaje (0-10)','Señales de CALIDAD (0-3) + señales de PRECIO a favor (0-4) + señales de SALUD (0-3). Ojo: la categoría manda sobre el número — una Neutral con 9 sigue siendo Neutral.')
    R=FILA(R,'Calidad','Tres lentes: Greenblatt (ROIC ≥20% actual y a 5 años) · MSCI (ROE ≥15% con deuda sana) · AQR (rentable, crece, segura, recompra). EXCELENTE = los 3 en Sí · BUENA = 2 · DÉBIL = 0-1. La calidad es la puerta: sin EXCELENTE no hay Deep Dive ni Watchlist.')
    R=FILA(R,'Precio','Cuatro jueces: EV/EBIT ≤10 (y positivo) · descuento ≥20% frente al Fair Value · FCF Yield ≥4% · la etiqueta del consenso de analistas. MUY BARATA = 3-4 a favor · BARATA = 2 · MIXTA = 1 o empate · CARA = más en contra que a favor. El veredicto siempre es neto.')
    R=FILA(R,'Salud financiera','Tres señales: Piotroski ≥7 · Altman ≥3 · M-Score sin alerta. EXCELENTE = 3 · MIXTA = 2 · DÉBIL = 0-1 o Altman <1.81. Salud DÉBIL saca a la empresa de las listas de acción.')
    R+=1
    R=SEC(R,'06','CÓMO LEERLO',TE)
    R=FILA(R,'Solo números','Todo este archivo es análisis cuantitativo: no ve marcas, ventajas competitivas, disrupciones ni calidad de la gerencia. Por eso una empresa con 10/10 puede ser una mala inversión, y por eso el paso siguiente siempre es el análisis cualitativo.')
    R=FILA(R,'¿Roja y barata?','Los cuatro jueces de precio pueden discrepar, y esa discrepancia es información. El veredicto es por mayoría neta, y la celda roja te dice cuál juez discrepó — ese es el primer punto a investigar.')
    R=FILA(R,'Colores','Bandas de columnas: VERDE = calidad · DORADO = precio · AZUL = salud. Encabezado oscuro = columna calculada por el análisis; claro = tu columna original de InvestingPro. En celdas: ROJO = falló su umbral · ÁMBAR = precaución · VERDE = a favor. Pasa el mouse por los encabezados para ver el umbral exacto de cada indicador.')
    R=FILA(R,'Los datos caducan','Precios y métricas son una foto del día del export. Correr el filtro cada mes dará resultados distintos — y eso es lo esperado.')
    R+=1
    R=SEC(R,'07','EDUCATIVO, NO ASESORÍA',CO)
    R=TXT(R,'Este archivo es una herramienta educativa de filtrado. No constituye asesoría financiera ni una recomendación de inversión. Ninguna categoría es una orden de comprar o vender: "Deep Dive" significa investigar a fondo, no comprar. Los umbrales son referencias de industria, ajustables. La decisión final requiere tu propio análisis y tu situación personal.')
    R+=1
    _cl(R,2,'INVESTINGPRO',color=NV,bold=True,size=9,fill=AM,align='center')
    _lk=_cl(R,3,'  Consigue InvestingPro con el descuento exclusivo de la comunidad → https://ainvertir.org/InvestingPro25',color=AM,bold=True,size=10,fill=CD)
    _lk.hyperlink='https://ainvertir.org/InvestingPro25'; ins.row_dimensions[R].height=20; R+=1
    _cl(R,3,'Transparencia: es el enlace de afiliado del canal — recomendamos solo lo que usamos.',color=MU,size=9,ital=True,fill=CD)
    s.sheet_properties.tabColor='F59C00'; ins.sheet_properties.tabColor='003850'
    wb.active=0
    wb.save(outpath)
    _fix_databar(str(outpath), rng_barras)
    return cts, funds, sectx, exc, tr, len(noF)+sectx

# ───────────────────────── 4 · DASHBOARD INTERACTIVO 2.0 ─────────────────────────

def _fix_databar(path, sqref):
    """Reemplaza la barra de datos básica por la versión moderna de Excel (extensión x14):
    eje automático en cero, positivos en verde a la derecha, negativos en ROJO a la izquierda.
    openpyxl no expone esta extensión, así que se inyecta en el XML del archivo."""
    import zipfile, re, shutil, tempfile
    GUID = '{DA7ABA51-0000-4000-8000-AA1DEC0DE001}'
    legacy = (f'<conditionalFormatting sqref="{sqref}"><cfRule type="dataBar" priority="1">'
              f'<dataBar><cfvo type="min"/><cfvo type="max"/><color rgb="FF63C384"/></dataBar>'
              f'<extLst><ext uri="{{B025F937-C7B1-47D3-B67F-A62EFF666E3E}}" '
              f'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
              f'<x14:id>{GUID}</x14:id></ext></extLst></cfRule></conditionalFormatting>')
    x14 = (f'<extLst><ext uri="{{78C0D931-6437-407d-A8EE-F0AAD7539E65}}" '
           f'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
           f'<x14:conditionalFormattings><x14:conditionalFormatting '
           f'xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
           f'<x14:cfRule type="dataBar" id="{GUID}">'
           f'<x14:dataBar minLength="0" maxLength="100" gradient="1" border="1" '
           f'negativeBarColorSameAsPositive="0" negativeBarBorderColorSameAsPositive="0" axisPosition="automatic">'
           f'<x14:cfvo type="autoMin"/><x14:cfvo type="autoMax"/>'
           f'<x14:borderColor rgb="FF63C384"/><x14:negativeFillColor rgb="FFFF0000"/>'
           f'<x14:negativeBorderColor rgb="FFFF0000"/><x14:axisColor rgb="FF000000"/>'
           f'</x14:dataBar><xm:sqref>{sqref}</xm:sqref></x14:cfRule>'
           f'</x14:conditionalFormatting></x14:conditionalFormattings></ext></extLst></worksheet>')
    tmp = tempfile.mktemp(suffix='.xlsx')
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item.startswith('xl/worksheets/') and item.endswith('.xml') and b'<dataBar' in data:
                xml = data.decode('utf-8')
                xml = re.sub(r'<conditionalFormatting[^>]*>\s*<cfRule type="dataBar".*?</conditionalFormatting>',
                             legacy, xml, count=1, flags=re.S)
                xml = xml.replace('</worksheet>', x14)
                data = xml.encode('utf-8')
            zout.writestr(item, data)
    shutil.move(tmp, path)

def _fmt(v, kind='num'):
    if v is None: return 'N/D'
    if kind == 'pct': return f'{v:.1f}%'
    if kind == 'up': return f'{v*100:+.1f}%'
    if kind == 'int': return f'{int(v)}'
    return f'{v:.2f}'

def _chip_class(v):
    m = {'MUY BARATA':'c-teal','BARATA':'c-teal2','EXCELENTE':'c-teal','MIXTA':'c-amber','CARA':'c-coral',
         'Sí':'c-teal','Medio':'c-amber','No':'c-coral','N/D':'c-mute','—':'c-mute','DÉBIL':'c-coral'}
    return m.get(v, 'c-amber')

DASH_CSS = """
:root{--navy:#0A1B2D;--petrol:#003850;--amber:#F59C00;--orange:#F1801B;--teal:#6DADAA;--coral:#E07068;--w:#FFFFFF;--mut:rgba(255,255,255,.55);--line:rgba(255,255,255,.08)}
*{margin:0;padding:0;box-sizing:border-box}
body{background:radial-gradient(1200px 500px at 70% -10%, rgba(0,56,80,.55), transparent 60%), var(--navy);color:var(--w);font-family:Montserrat,'Segoe UI','Helvetica Neue',Arial,sans-serif;padding:34px 22px 46px}
.wrap{max-width:1140px;margin:0 auto}
header{display:flex;justify-content:space-between;align-items:center;gap:18px;flex-wrap:wrap;margin-bottom:20px}
header img{height:40px}
.meta{text-align:right;font-size:12px;color:var(--mut);line-height:1.55}
.meta b{color:var(--teal);font-style:italic;font-weight:600}
h1{font-weight:900;text-transform:uppercase;font-size:26px;letter-spacing:.06em;margin:2px 0 18px}
h1 span{color:var(--amber)}
.tabs{display:flex;gap:10px;margin-bottom:22px}
.tab{background:transparent;border:1px solid var(--line);color:var(--mut);border-radius:99px;padding:9px 18px;font-family:inherit;font-weight:800;font-size:11px;letter-spacing:.12em;text-transform:uppercase;cursor:pointer}
.tab.on{background:var(--amber);border-color:var(--amber);color:var(--navy)}
.funnel{display:flex;align-items:stretch;gap:14px;flex-wrap:wrap;margin-bottom:26px}
.stage{flex:1;min-width:180px;background:var(--petrol);border:1px solid var(--line);border-radius:16px;padding:20px 22px}
.stage .n{font-family:'Playfair Display',Georgia,'Times New Roman',serif;font-weight:800;font-size:52px;line-height:1}
.stage .l{margin-top:8px;font-size:11.5px;letter-spacing:.14em;text-transform:uppercase;color:var(--mut);font-weight:600}
.stage.hl{border-color:var(--amber);box-shadow:0 0 0 1px var(--amber), 0 10px 34px rgba(245,156,0,.18)}
.stage.hl .n{color:var(--amber)}
.arrow{align-self:center;color:var(--amber);font-size:30px;font-weight:900}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px;margin-bottom:34px}
.kpi{background:var(--petrol);border:1px solid var(--line);border-left:4px solid var(--c);border-radius:14px;padding:16px 18px}
.kpi .v{font-size:34px;font-weight:800}
.kpi .v .ke{font-size:22px;margin-right:9px;vertical-align:2px}
.kpi .t{font-size:11px;letter-spacing:.14em;text-transform:uppercase;color:var(--mut);font-weight:700;margin-top:2px}
.kpi .s{font-size:11.5px;color:var(--coral);margin-top:6px}
.sect{display:flex;align-items:center;gap:10px;margin:6px 0 14px}
.sect i{width:10px;height:10px;border-radius:50%;background:var(--amber);display:inline-block;flex:none}
.sect h2{font-size:14px;font-weight:800;letter-spacing:.16em;text-transform:uppercase}
.sect small{color:var(--mut);font-weight:600;font-size:12px;margin-left:auto}
.card{background:rgba(0,56,80,.45);border:1px solid var(--line);border-radius:16px;padding:8px 14px;margin-bottom:30px}
.grid7{display:grid;grid-template-columns:34px 1.35fr 1.15fr 112px 112px 122px 1fr;gap:12px;align-items:center}
.thead{padding:12px 8px 9px;border-bottom:2px solid rgba(255,255,255,.14)}
.thead div{font-size:9.5px;letter-spacing:.15em;text-transform:uppercase;color:var(--mut);font-weight:800;line-height:1.35}
.ddrow{border-bottom:1px solid var(--line);cursor:pointer}
.ddrow:last-child{border-bottom:none}
.ddrow .main{padding:11px 30px 11px 8px;position:relative}
.chev{position:absolute;right:10px;top:50%;transform:translateY(-50%);color:var(--mut);font-size:11px;transition:transform .18s,color .18s}
.ddrow.open .chev{transform:translateY(-50%) rotate(180deg);color:var(--amber)}
.ddrow:hover .chev{color:var(--amber)}
.rk{color:var(--mut);font-weight:700;font-size:12px}
.emp b{font-size:15px;letter-spacing:.02em}
.tlink{color:inherit;text-decoration:none;border-bottom:1px dashed rgba(245,156,0,.55)}
.tlink:hover b{color:var(--amber)}
.emp span{display:block;font-size:11.5px;color:var(--mut)}
.sc{display:flex;align-items:center;gap:10px}
.bar{flex:1;height:7px;border-radius:99px;background:rgba(255,255,255,.09);overflow:hidden}
.bar i{display:block;height:100%;border-radius:99px;background:linear-gradient(90deg,var(--orange),var(--amber))}
.sc em{font-style:normal;font-weight:800;font-size:15px;min-width:44px;text-align:right}
.sc small{color:var(--mut);font-weight:600}
.chip{display:inline-block;padding:4px 10px;border-radius:99px;font-size:10.5px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;border:1px solid}
.c-teal{color:var(--teal);border-color:rgba(109,173,170,.6);background:rgba(109,173,170,.12)}
.c-teal2{color:var(--teal);border-color:rgba(109,173,170,.35);background:rgba(109,173,170,.06)}
.c-amber{color:var(--amber);border-color:rgba(245,156,0,.5);background:rgba(245,156,0,.1)}
.c-coral{color:var(--coral);border-color:rgba(224,112,104,.55);background:rgba(224,112,104,.12)}
.c-mute{color:var(--mut);border-color:transparent;background:transparent;text-transform:none}
.det{display:none;padding:4px 8px 16px 46px}
.ddrow.open .det{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:12px}
.ddrow.open{background:rgba(255,255,255,.025)}
.tcard{background:rgba(10,27,45,.55);border:1px solid var(--line);border-radius:12px;padding:12px 14px}
.tcard h4{font-size:10px;letter-spacing:.14em;text-transform:uppercase;color:var(--mut);margin-bottom:8px}
.tcard .vr{margin-bottom:8px}
.mets{display:flex;flex-wrap:wrap;gap:5px}
.met{font-size:10.5px;color:rgba(255,255,255,.75);background:rgba(255,255,255,.05);border:1px solid var(--line);border-radius:8px;padding:3px 7px}
.met b{color:var(--w);font-weight:700}
.hint{font-size:10.5px;color:var(--mut);text-align:right;padding:2px 8px 8px}
.warn{background:#C00000;color:#fff;border-radius:10px;padding:14px 18px;margin:0 0 22px;font-size:13.5px;line-height:1.5}
.warn .wcols{font-family:ui-monospace,Menlo,Consolas,monospace;font-size:12.5px;background:rgba(0,0,0,.22);padding:2px 7px;border-radius:5px}
.warn .wsub{opacity:.92;font-size:12.5px}
.wgrid{display:flex;flex-wrap:wrap;gap:8px;padding:12px 6px 16px}
.wchip{border:1px solid rgba(109,173,170,.35);color:#cfe3e2;border-radius:10px;padding:6px 10px;font-size:12px;font-weight:700}
.wchip i{font-style:normal;color:var(--teal);font-weight:800;margin-left:6px;font-size:11px}
a.wchip{text-decoration:none;cursor:pointer;transition:border-color .15s,color .15s,background .15s}
a.wchip:hover{border-color:rgba(245,156,0,.75);color:var(--amber);background:rgba(245,156,0,.07)}
a.wchip:hover i{color:var(--amber)}
footer{border-top:1px solid var(--line);padding-top:16px;font-size:11.5px;color:var(--mut);line-height:1.7}
footer .src{color:var(--teal);font-style:italic;margin-top:6px}
.gsec{background:rgba(0,56,80,.45);border:1px solid var(--line);border-radius:16px;padding:20px 24px;margin-bottom:18px}
.gsec h3{font-size:13px;font-weight:800;letter-spacing:.14em;text-transform:uppercase;color:var(--amber);margin-bottom:10px}
.gsec p{font-size:13.5px;line-height:1.75;color:rgba(255,255,255,.85);margin-bottom:8px}
.gsec .cat{display:flex;gap:12px;align-items:flex-start;margin-bottom:10px}
.gsec .cat .chip{flex:none;margin-top:2px}
.gsec ol{margin:4px 0 4px 20px}
.gsec li{font-size:13.5px;line-height:1.8;color:rgba(255,255,255,.85)}
.btnrow{display:flex;gap:10px;justify-content:flex-end;margin:-6px 0 16px}
.btn{background:transparent;border:1px solid var(--line);color:var(--mut);border-radius:10px;padding:7px 14px;font-family:inherit;font-weight:700;font-size:11px;cursor:pointer}
@media(max-width:820px){.grid7{grid-template-columns:26px 1.3fr 1fr}.thead div:nth-child(n+4){display:none}.main div:nth-child(n+4){grid-column:2/4}.det{padding-left:8px}}
@media print{body{background:#fff;color:#000}.tabs,.btnrow,.hint{display:none}#guia{display:block!important}.card,.gsec,.kpi,.stage{border-color:#ccc;background:#fff}.stage .n{color:#000}.meta b,.src{color:#555}}
"""

def generar_dashboard(evals, orden, outpath, export_name, fecha, cts, funds, exc, tr, analizadas, faltan=None):
    _fl = faltan or []
    _fl_nuc = [h for h in _fl if h in COLS_NUCLEO]
    _banner = ('<div class="warn"><b>⚠️ Faltan ' + str(len(_fl)) + ' indicadores en tu export</b> — agrégalos en el screener de InvestingPro y vuelve a exportar: <span class="wcols">'
               + ' · '.join(_h(x) for x in _fl) + '</span><br><span class="wsub">'
               + ('Sin ellos varias empresas quedan clasificadas por debajo de lo que les corresponde: esta corrida no es comparable con una completa.'
                  if _fl_nuc else 'No cambian la clasificación, pero pierdes contexto en el análisis.')
               + '</span></div>') if _fl else ''
    _sx = sum(1 for k in evals if evals[k].get('sectx'))
    _sx_note = ''
    _sx_kpi = (f'<div class="kpi" style="--c:rgba(255,255,255,.18)"><div class="v"><span class="ke">⚙️</span>{_sx}</div>'
               f'<div class="t">Omitidas por método</div><div class="s" style="color:var(--mut)">Financieras/Utilities — se analizan con otro criterio</div></div>') if _sx else ''
    logo_b64 = ''
    try:
        logo_b64 = base64.b64encode((ASSETS / 'logo_completo_blanco.png').read_bytes()).decode()
    except Exception:
        pass
    dd = [evals[k] for k in orden if not evals[k]['fund'] and evals[k]['balde'] == 'Deep Dive']
    # El Watchlist se ordena por capitalización de mercado (mayor → menor), no alfabéticamente:
    # las empresas sin dato de market cap van al final, desempatando por ticker.
    wl = sorted([evals[k] for k in orden if not evals[k]['fund'] and evals[k]['balde'] == 'Watchlist'],
                key=lambda e: (0, -e['mets']['mcap']) if e['mets'].get('mcap') is not None else (1, 0))
    rows = ''
    for i, e in enumerate(dd, 1):
        m = e['mets']
        lentes = e['L']
        moat = ''
        if m['roic'] is not None and m['roic5'] is not None and m['roic'] < m['roic5']:
            moat = '<div class="vr"><span class="chip c-amber">↘ ROIC actual bajo su promedio 5a — revisar moat</span></div>'
        tk = f'<a class="tlink" href="https://www.investing.com/pro/{_h(m["ft"])}" target="_blank" rel="noopener"><b>{_h(e["ticker"])}</b></a>' if m.get('ft') else f'<b>{_h(e["ticker"])}</b>'
        det = f'''<div class="det">
<div class="tcard"><h4>Tribunal de calidad — {e['n_cal']}/3 lentes</h4>
<div class="vr"><span class="chip {_chip_class(lentes[0])}">Greenblatt: {lentes[0]}</span> <span class="chip {_chip_class(lentes[1])}">MSCI: {lentes[1]}</span> <span class="chip {_chip_class(lentes[2])}">AQR: {lentes[2]}</span></div>{moat}
<div class="mets"><span class="met">ROIC <b>{_fmt(m['roic'],'pct')}</b></span><span class="met">ROIC 5a <b>{_fmt(m['roic5'],'pct')}</b></span><span class="met">ROE <b>{_fmt(m['roe'],'pct')}</b></span><span class="met">Margen bruto <b>{_fmt(m['gm'],'pct')}</b></span><span class="met">BPA 5a <b>{_fmt(m['epsg'],'pct')}</b></span><span class="met">Ventas 5a <b>{_fmt(m['revg'],'pct')}</b></span></div></div>
<div class="tcard"><h4>Salud financiera — {e['n_sal']}/3 señales</h4>
<div class="vr"><span class="chip {_chip_class(e['salud'])}">{e['salud']}</span></div>
<div class="mets"><span class="met">Piotroski <b>{_fmt(m['pio'],'int')}/9</b></span><span class="met">Altman Z <b>{_fmt(m['alt'])}</b></span><span class="met">Beneish M <b>{_fmt(m['m'])}</b></span><span class="met">Deuda/Capital <b>{_fmt(m['dc'],'pct')}</b></span></div></div>
<div class="tcard"><h4>Tribunal de precio — {e['n_val']} a favor · {e['v_c']} en contra</h4>
<div class="vr"><span class="chip {_chip_class(e['precio'])}">{e['precio']}</span></div>
<div class="mets"><span class="met">EV/EBIT <b>{_fmt(m['evebit'])}</b></span><span class="met">FCF Yield <b>{_fmt(m['fcfy'],'pct')}</b></span><span class="met">Precio vs Fair Value <b>{_fmt(e['upside'],'up') if e['upside'] is not None else 'N/D'}</b></span><span class="met">Analistas <b>{(m['lbl'] or 'N/D').title()}</b></span></div></div>
</div>'''
        nota = e['alertas'] if e['alertas'] != '—' else '—'
        rows += f'''<div class="ddrow"><div class="grid7 main"><div class="rk">{i:02d}</div><div class="emp">{tk}<span>{_h(e['name'][:34])}</span></div><div class="sc"><div class="bar"><i style="width:{e['punt']*10}%"></i></div><em>{e['punt']}<small>/10</small></em></div><div><span class="chip {_chip_class(e['cal'])}">{e['cal']}</span></div><div><span class="chip {_chip_class(e['salud'])}">{e['salud']}</span></div><div><span class="chip {_chip_class(e['precio'])}">{e['precio']}</span></div><div><span class="chip {'c-mute' if nota=='—' else ('c-coral' if 'M-Score' in nota else 'c-amber')}">{nota}</span></div><span class="chev">▾</span></div>{det}</div>\n'''
    def _wchip(e):
        _in = f'{_h(e["ticker"])}<i>{e["punt"]}</i>'
        _ft = e['mets'].get('ft')
        return (f'<a class="wchip" href="https://www.investing.com/pro/{_h(_ft)}" target="_blank" rel="noopener">{_in}</a>'
                if _ft else f'<span class="wchip">{_in}</span>')
    wlh = ''.join(_wchip(e) for e in wl)
    html = f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Screening Cuantitativo — Aprende a Invertir</title>
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;800;900&family=Playfair+Display:wght@700;800&display=swap" rel="stylesheet">
<style>{DASH_CSS}</style></head><body><div class="wrap">
<header><img src="data:image/png;base64,{logo_b64}" alt="Aprende a Invertir">
<div class="meta">INFORME DE SCREENING CUANTITATIVO<br><b>Export: {export_name} · Corrida: {fecha} · {MOTOR_VERSION}</b>{_sx_note}</div></header>
{_banner}
<h1>De {analizadas:,} acciones a <span>{cts['Deep Dive']} candidatas</span></h1>
<div class="tabs"><button class="tab on" onclick="tab('resumen',this)">Resumen</button><button class="tab" onclick="tab('guia',this)">Cómo leer este informe</button></div>
<section id="resumen">
<div class="funnel">
<div class="stage"><div class="n">{analizadas:,}</div><div class="l">Empresas analizadas</div></div>
<div class="arrow">→</div>
<div class="stage"><div class="n">{exc}</div><div class="l">Calidad y salud sólidas</div></div>
<div class="arrow">→</div>
<div class="stage hl"><div class="n">{cts['Deep Dive']}</div><div class="l">Deep Dive · excelentes y a buen precio</div></div>
</div>
<div class="kpis">
<div class="kpi" style="--c:var(--amber)"><div class="v"><span class="ke">🔬</span>{cts['Deep Dive']}</div><div class="t">Deep Dive</div></div>
<div class="kpi" style="--c:var(--teal)"><div class="v"><span class="ke">📋</span>{cts['Watchlist']}</div><div class="t">Watchlist</div></div>
<div class="kpi" style="--c:rgba(255,255,255,.35)"><div class="v"><span class="ke">⚪</span>{cts['Neutral']}</div><div class="t">Neutral</div></div>
<div class="kpi" style="--c:var(--coral)"><div class="v"><span class="ke">❌</span>{cts['Descartada']}</div><div class="t">Descartadas</div><div class="s">{tr} posibles trampas de valor</div></div>{_sx_kpi}
</div>
<div class="sect"><i></i><h2>Deep Dive — investigar a fondo</h2><small>no es una lista de compra</small></div>
<div class="card">
<div class="grid7 thead"><div>#</div><div>Empresa</div><div>Puntaje (0-10)</div><div>Calidad</div><div>Salud financiera</div><div>Precio</div><div>Nota</div></div>
{rows}<div class="hint">clic en una fila para ver sus tres tribunales · clic en el ticker para abrir su ficha en InvestingPro</div>
</div>
<div class="sect"><i style="background:var(--teal)"></i><h2>Watchlist — excelentes esperando su precio</h2><small>{cts['Watchlist']} empresas</small></div>
<div class="card"><div class="wgrid">{wlh}</div><div class="hint">clic en un ticker para abrir su ficha en InvestingPro</div></div>
</section>
<section id="guia" style="display:none">
<div class="btnrow"><button class="btn" onclick="window.print()">Imprimir / guardar PDF</button></div>
<div class="gsec"><h3>La filosofía — por qué este orden</h3><p>"Es mucho mejor comprar una empresa maravillosa a un precio justo que una empresa justa a un precio maravilloso" — Warren Buffett, 1989. El filtro aplica esa idea en orden: primero la <b>CALIDAD</b> como puerta eliminatoria, la <b>SALUD FINANCIERA</b> como cinturón de seguridad (anti-quiebra y anti-manipulación contable), y al final el <b>PRECIO</b> decide la lista. Es la evolución del value investing de Buffett y Munger — calidad primero — no el enfoque clásico de comprar cualquier cosa barata.</p></div>
<div class="gsec"><h3>Cómo leer cada categoría</h3>
<div class="cat"><span class="chip c-teal">Deep Dive</span><p>Excelente empresa a excelente precio — según los números. El nombre es literal: investigar a fondo antes de cualquier decisión. Los números son la mitad de la tarea.</p></div>
<div class="cat"><span class="chip c-teal2">Watchlist</span><p>Excelente empresa cuyo precio aún no es atractivo. Va al radar, esperando un mejor precio.</p></div>
<div class="cat"><span class="chip c-amber">Neutral</span><p>Ni excelente ni mala según los datos de hoy. Puede graduarse si su calidad mejora.</p></div>
<div class="cat"><span class="chip c-coral">Descartada</span><p>Hoy, ni la calidad ni el precio son atractivos. NO es un pronóstico: una descartada puede subir — el filtro no predice precios, filtra negocios por sus números.</p></div>
<div class="cat"><span class="chip c-amber">Posible trampa de valor</span><p>Descartada que además se ve barata. Lo barato es el anzuelo: cuando una empresa débil está muy descontada, el descuento suele tener dueño.</p></div></div>
<div class="gsec"><h3>El puntaje y los tres tribunales</h3>
<p><b>Puntaje (0-10)</b> = señales de calidad (0-3) + señales de precio a favor (0-4) + señales de salud (0-3).</p>
<p><b>Calidad — 3 lentes:</b> Greenblatt (ROIC ≥20% actual y a 5 años) · MSCI (ROE ≥15% con deuda sana) · AQR (rentable, crece, segura, recompra). EXCELENTE = los 3 en Sí. Sin EXCELENTE no hay Deep Dive ni Watchlist.</p>
<p><b>Precio — 4 jueces, veredicto neto:</b> EV/EBIT ≤10 (y positivo) · descuento ≥20% vs Fair Value (promedio de 14 modelos) · FCF Yield ≥4% · etiqueta del consenso de analistas. Las señales en contra también cuentan.</p>
<p><b>Salud financiera — 3 señales:</b> Piotroski ≥7 · Altman ≥3 · M-Score sin alerta. Altman &lt;1.81 = salud DÉBIL automática, y la salud DÉBIL saca a la empresa de las listas de acción.</p></div>
<div class="gsec"><h3>El checklist del Deep Dive — la mitad humana</h3>
<p>El filtro cubre los números. Antes de cualquier decisión, responde esto sobre cada candidata:</p>
<ol><li>¿Cómo gana dinero exactamente? Explícalo en una frase simple.</li>
<li>¿Qué la protege de la competencia? (la ventaja competitiva o "moat")</li>
<li>¿POR QUÉ está barata? El mercado rara vez regala calidad — esta es la pregunta anti-trampa.</li>
<li>¿Qué podría romper el negocio en 5-10 años? (disrupción, tecnología, regulación)</li>
<li>¿Los números vienen mejorando, o solo son buenos hoy? Empieza por el ROIC: si el actual está por debajo de su promedio de 5 años, pregunta por qué — puede ser la primera señal de erosión de la ventaja competitiva. Los números pueden seguir fuertes años después de que el negocio se rompió.</li>
<li>Abre la ficha de la empresa: mira el RANGO de los 14 modelos de Fair Value. Rango ancho = valuación frágil, exige más descuento.</li>
<li>Si tiene nota de M-Score: revisa la contabilidad con lupa antes de seguir.</li>
<li>¿Entiendes tú el negocio? Si no está en tu círculo de competencia, pasa a la siguiente.</li></ol>
<p>Si una respuesta no te convence, descártala y sigue — el filtro te dará más candidatas el próximo mes.</p></div>
<div class="gsec"><h3>Los datos caducan</h3><p>Precios y métricas son una foto del día del export. Correr el filtro de nuevo (por ejemplo, cada mes) dará resultados distintos — y eso es lo esperado: las oportunidades aparecen y desaparecen.</p></div>
<div class="gsec"><h3>Educativo, no asesoría</h3><p>Esta herramienta es educativa. No constituye asesoría financiera ni recomendación de inversión. Ninguna categoría es una orden de comprar o vender. Los umbrales son referencias de industria, ajustables según tu criterio. La decisión final requiere tu propio análisis y tu situación personal.</p></div>
</section>
<footer>Herramienta educativa de filtrado — no constituye asesoría financiera ni recomendación de inversión. Ninguna categoría es una orden de comprar o vender: "Deep Dive" significa investigar a fondo, no comprar. Este filtro no predice precios. Los números son la mitad de la tarea.
<div class="src">Fuente de datos: InvestingPro · Método: principios documentados de Greenblatt, Piotroski, Altman, Beneish, MSCI Quality y AQR · Skill Aprende a Invertir · {MOTOR_VERSION} · corrida {fecha}</div></footer>
</div>
<script>
function tab(id,btn){{document.getElementById('resumen').style.display=id==='resumen'?'block':'none';document.getElementById('guia').style.display=id==='guia'?'block':'none';document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));btn.classList.add('on');}}
document.querySelectorAll('.ddrow').forEach(r=>r.addEventListener('click',ev=>{{if(ev.target.closest('a'))return;r.classList.toggle('open');}}));
</script></body></html>'''
    Path(outpath).write_text(html, encoding='utf-8')

# ───────────────────────── 5 · INFORME EJECUTIVO + MAIN ─────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('export')
    ap.add_argument('--outdir', default='/mnt/user-data/outputs')
    args = ap.parse_args()
    fecha = date.today().isoformat()
    export_name = Path(args.export).name
    outdir = Path(args.outdir); outdir.mkdir(parents=True, exist_ok=True)

    filas, hmap, faltantes = leer_export(args.export)
    sospechosas = verificar_convencion(filas, hmap)
    faltan_cols = [c for c in COLS_NUCLEO + COLS_OPCIONALES if c not in hmap]
    raw, num, pctv = _mk_accessors(hmap)
    evals = {k: evaluate(row, raw, num, pctv) for k, row in enumerate(filas)}
    ob = {'Deep Dive': 0, 'Watchlist': 1, 'Neutral': 2, 'Descartada': 3}
    orden = sorted(evals.keys(), key=lambda k: (9, 0, '') if evals[k]['fund']
                   else (ob[evals[k]['balde']], -evals[k]['punt'], evals[k]['ticker']))

    xlsx = outdir / f'Screening_AAI_{fecha}.xlsx'
    cts, funds, sectx, exc, tr, analizadas = generar_excel(filas, hmap, evals, orden, xlsx, export_name, fecha)
    dash = outdir / f'Dashboard_AAI_{fecha}.html'
    generar_dashboard(evals, orden, dash, export_name, fecha, cts, funds, exc, tr, analizadas, faltan=faltan_cols)

    dd = [evals[k] for k in orden if not evals[k]['fund'] and evals[k]['balde'] == 'Deep Dive']
    print('══ INFORME EJECUTIVO · Screening Aprende a Invertir ══')
    if faltantes:
        print(f'AVISO — faltan {len(faltantes)} indicadores núcleo en tu export: {", ".join(faltantes)}.')
        print('Agrégalos en el screener de InvestingPro (nombre exacto) y vuelve a exportar. Sin ellos, varias empresas')
        print('quedan clasificadas por debajo de lo que les corresponde: esta corrida no es comparable con una completa.')
        print('En el Excel y el dashboard verás marcado en rojo exactamente qué falta.')
    if 'Full Ticker' not in hmap:
        print('AVISO — falta la columna "Full Ticker": el informe se genera igual, pero los tickers '
              'del dashboard NO quedan enlazados a InvestingPro. Agrégala al screener para recuperar los enlaces.')
    if sospechosas:
        print(f'AVISO — estas columnas no parecen venir en decimales: {", ".join(sospechosas)}. '
              'Si InvestingPro cambió el formato del export, los umbrales quedarían mal calibrados: '
              'reporta este aviso antes de usar los resultados.')
    print(f'Export: {export_name} · Corrida: {fecha} · {MOTOR_VERSION}')
    print(f'Embudo: {analizadas} empresas analizadas → {exc} con calidad excelente y salud sólida → {cts["Deep Dive"]} en Deep Dive')
    print(f'Categorías: Deep Dive {cts["Deep Dive"]} · Watchlist {cts["Watchlist"]} · Neutral {cts["Neutral"]} · Descartadas {cts["Descartada"]} ({tr} con nota de posible trampa) · Fondos/ETFs excluidos {funds}' + (f' · Omitidas por método (Financieras/Utilities) {sectx}' if sectx else ''))
    print('── Deep Dive (excelente calidad + salud sana + buen precio) ──')
    for i, e in enumerate(dd, 1):
        nota = f' · {e["alertas"]}' if e['alertas'] != '—' else ''
        print(f'{i:2d}. {e["ticker"]:6s} {e["name"][:30]:30s} {e["punt"]}/10 · {e["precio"]} · salud {e["salud"]}{nota}')
    print(f'Archivos generados: {xlsx} · {dash}')
    print('Recordatorio: herramienta educativa; Deep Dive = investigar a fondo, no comprar. Los números son la mitad de la tarea.')

if __name__ == '__main__':
    main()
