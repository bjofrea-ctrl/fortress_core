"""
TEST END-TO-END — cierra la CLASE de bug de la Fase 4.

El bug auditado: los 25 tests de Fase 4 verificaban lo que el endpoint LEE
(screen_<date>.json, dashboard_<date>.html fabricado por fixtures del propio
test), pero NADIE verificaba que el JOB los produce. Un futuro cambio que
rompa el render del dashboard/Excel pasaría verde porque los tests no lo
miran.

Este archivo corre el JOB COMPLETO (runner + screen_payload + motor canónico
render) contra fixtures FMP reales (sin red) y verifica que produce los
artefactos que el endpoint necesita:

    screen_<date>.json          — leído por /screen/latest, /screen?date=
    dashboard_<date>.html       — leído por /screen/dashboard.html
    Screening_AAI_<date>.xlsx   — leído por /screen/export.xlsx

Si el job no genera el HTML, este test FALLA EN ROJO. Es el guard de la
clase completa, no de la instancia.
"""

import json
import os
import sys
from pathlib import Path

import pytest

from app.core.fundamentals_screen import screen_payload
from app.config import settings
from scripts import run_fundamentals_screen as job


# ---------------------------------------------------------------------------
# Fixtures FMP reales (Fase 1) para construir payloads sin red.
# ---------------------------------------------------------------------------
_FIXTURES = ["fmp_income_statement_aapl.json", "fmp_balance_sheet_aapl.json",
             "fmp_cash_flow_aapl.json", "fmp_profile_aapl.json",
             "fmp_price_target_aapl.json"]
_FMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures", "fmp")


def _load_fixture(name):
    with open(os.path.join(_FMP_DIR, name)) as f:
        return json.load(f)


def _first(x):
    return x[0] if isinstance(x, list) and x else x


def _payload_for(symbol):
    """Payload FMP de AAPL reutilizado por símbolo (los fixtures solo tienen
    AAPL real — para MSFT reutilizamos los mismos números; lo que importa en
    el E2E es el FLUJO completo, no la exactitud del ticker)."""
    payload = {
        "symbol": symbol,
        "income_statement": _load_fixture("fmp_income_statement_aapl.json"),
        "balance_sheet": _load_fixture("fmp_balance_sheet_aapl.json"),
        "cash_flow": _load_fixture("fmp_cash_flow_aapl.json"),
        "profile": _first(_load_fixture("fmp_profile_aapl.json")),
        "price_target_consensus": _first(_load_fixture("fmp_price_target_aapl.json")),
    }
    if symbol != "AAPL":
        payload["profile"]["symbol"] = symbol
    return payload


def test_end_to_end_job_produces_artifacts_que_usa_el_endpoint(monkeypatch, tmp_path):
    """El job completo (runner) produce los 3 artefactos del endpoint.

    Cierra la CLASE de bug: si un futuro cambio rompe el render del motor
    canónico, este test deja de encontrar dashboard_/Screening_ y falla.
    """
    monkeypatch.setattr(settings, "FMP_API_KEY", "fake-key")
    monkeypatch.setattr(job, "CACHE_DIR", str(tmp_path))
    monkeypatch.setattr(job, "STATE_PATH", str(tmp_path / "state.json"))
    monkeypatch.setattr(job, "DAILY_FMP_BUDGET", 240)
    monkeypatch.setattr(job, "BATCH_PAUSE_SECONDS", 0)

    class FakeFmp:
        is_available = lambda self: True

    class FakeIngester:
        fmp = FakeFmp()

        def ingest_symbol(self, sym):
            if sym == "BAD":
                return None  # ejercicio del path de fallo: no produce artefacto
            return _payload_for(sym)

    monkeypatch.setattr(job, "FundamentalsIngestion", lambda: FakeIngester())

    rc = job.main(["run_fundamentals_screen",
                   "--universe", "AAPL,MSFT,BAD",
                   "--date", "2026-08-28"])
    assert rc == 0, f"job rc={rc}; el job debe terminar 0 con artefactos OK"

    # --- Los 3 artefactos que el endpoint consume deben existir ---
    screen_json = tmp_path / "screen_2026-08-28.json"
    dashboard_html = tmp_path / "dashboard_2026-08-28.html"
    screening_xlsx = tmp_path / "Screening_AAI_2026-08-28.xlsx"
    assert screen_json.exists(), "el job no escribió screen_2026-08-28.json"
    assert dashboard_html.exists(), \
        f"EL DASHBOARD NO SE GENERÓ — el endpoint /screen/dashboard.html " \
        f"sirve 404 para siempre. Archivo esperado: {dashboard_html}"
    assert screening_xlsx.exists(), \
        f"EL EXCEL NO SE GENERÓ — /screen/export.xlsx no tendría qué servir: " \
        f"{screening_xlsx}"

    # --- Contenido de screen.json: los 2 OK, 1 BAD en failed ---
    artifact = json.loads(screen_json.read_text())
    assert artifact["date"] == "2026-08-28"
    assert set(artifact["results"].keys()) == {"AAPL", "MSFT"}
    assert "BAD" not in artifact["results"]
    assert artifact["completed_count"] == 2
    assert artifact["failed_count"] == 1

    # --- HTML: estructura del motor canónico, no un stub ---
    html = dashboard_html.read_text(encoding="utf-8")
    assert "<!DOCTYPE html>" in html
    assert "Screening Cuantitativo" in html            # título del motor canónico
    assert "Deep Dive" in html                          # categoría del motor
    assert "grid7" in html                              # layout del motor
    assert "data:image/png;base64" in html               # branding (logo del motor en base64)
    assert "header img" in html                          # CSS del logo del motor
    assert "corrida 2026-08-28" in html               # datos entraron (traza de corrida del motor)

    # --- XLSX: workbook real de openpyxl con las hojas del motor canónico ---
    import openpyxl
    wb = openpyxl.load_workbook(screening_xlsx, data_only=True, read_only=True)
    assert "Screening" in wb.sheetnames
    assert "Instructivo" in wb.sheetnames
    wb.close()
# ---------------------------------------------------------------------------
# Tests por artefacto individual — para aislar cuál rompe
# ---------------------------------------------------------------------------

def _run_job(monkeypatch, tmp_path, universe="AAPL"):
    monkeypatch.setattr(settings, "FMP_API_KEY", "fake-key")
    monkeypatch.setattr(job, "CACHE_DIR", str(tmp_path))
    monkeypatch.setattr(job, "STATE_PATH", str(tmp_path / "state.json"))
    monkeypatch.setattr(job, "DAILY_FMP_BUDGET", 240)
    monkeypatch.setattr(job, "BATCH_PAUSE_SECONDS", 0)

    class FakeFmp:
        is_available = lambda self: True

    class FakeIngester:
        fmp = FakeFmp()
        def ingest_symbol(self, sym):
            return _payload_for(sym)

    monkeypatch.setattr(job, "FundamentalsIngestion", lambda: FakeIngester())
    return job.main(["run_fundamentals_screen", "--universe", universe,
                     "--date", "2026-08-28"])


def test_end_to_end_job_genera_dashboard_html(monkeypatch, tmp_path):
    rc = _run_job(monkeypatch, tmp_path)
    assert rc == 0
    html_path = tmp_path / "dashboard_2026-08-28.html"
    assert html_path.exists()
    html = html_path.read_text(encoding="utf-8")
    assert "grid7" in html, "el HTML del motor canónico usa el layout grid7"


def test_end_to_end_job_genera_excel(monkeypatch, tmp_path):
    rc = _run_job(monkeypatch, tmp_path)
    assert rc == 0
    xlsx = tmp_path / "Screening_AAI_2026-08-28.xlsx"
    assert xlsx.exists(), "el job debe generar Screening_AAI_<fecha>.xlsx"
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    assert "Screening" in wb.sheetnames
    assert "Instructivo" in wb.sheetnames
    wb.close()


def test_end_to_end_job_rc3_si_render_falla(monkeypatch, tmp_path):
    """Si el render del dashboard falla (motor canónico), el job debe
    devolver rc=3 — NO 0 con el dashboard perdido en silencio."""
    monkeypatch.setattr(settings, "FMP_API_KEY", "fake-key")
    monkeypatch.setattr(job, "CACHE_DIR", str(tmp_path))
    monkeypatch.setattr(job, "STATE_PATH", str(tmp_path / "state.json"))
    monkeypatch.setattr(job, "DAILY_FMP_BUDGET", 240)
    monkeypatch.setattr(job, "BATCH_PAUSE_SECONDS", 0)

    class FakeFmp:
        is_available = lambda self: True

    class FakeIngester:
        fmp = FakeFmp()
        def ingest_symbol(self, sym):
            return _payload_for(sym)

    monkeypatch.setattr(job, "FundamentalsIngestion", lambda: FakeIngester())

    # Rompemos el render para verificar que el job detecta el fallo.
    def _boom(results, run_date, outdir):
        raise RuntimeError("motor canónico simulado roto")
    import app.core.fundamentals_artifacts as fa
    monkeypatch.setattr(fa, "render_artifacts", _boom)

    rc = job.main(["run_fundamentals_screen", "--universe", "AAPL",
                   "--date", "2026-08-28"])
    assert rc == 3, "render fallido debe reportar rc=3, no 0"
    # El state.json NO debe marcar la corrida como exitosa
    state = json.loads((tmp_path / "state.json").read_text())
    assert state.get("last_run_finished") is None