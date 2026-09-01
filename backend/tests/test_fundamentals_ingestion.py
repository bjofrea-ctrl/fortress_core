"""Tests de Fase 1 — ingesta de datos crudos de fundamentales.

Cubren `FmpClient` y `FundamentalsIngestion` con fixtures JSON y mocks: NUNCA
se toca la red real en esta suite (se pincha `FmpClient._fetch` y se simula el
cruce Finnhub). Replican el estándar de `test_data_ingestion.py`
(monkeypatch + tmp_path + capsys) y las señales de log distintas por rama del
esquema corregido (>=1): cache miss / cache vacío-trunco / cache fresco /
cache stale → refresh / refresh vacío conserva cache.
"""
import json
import os
import time

import pytest

from app.core.fundamentals_ingestion import (
    FmpClient,
    FundamentalsIngestion,
    TTL_DAYS,
)

FIXT = os.path.join(os.path.dirname(__file__), "fixtures", "fmp")


def _load(name):
    with open(os.path.join(FIXT, name)) as f:
        return json.load(f)


def _fmp_responses():
    """Mapa endpoint -> fixture para la inyección (sin tocar red)."""
    return {
        "income-statement/AAPL": _load("fmp_income_statement_aapl.json"),
        "balance-sheet-statement/AAPL": _load("fmp_balance_sheet_aapl.json"),
        "cash-flow-statement/AAPL": _load("fmp_cash_flow_aapl.json"),
        "profile/AAPL": _load("fmp_profile_aapl.json"),
        "price-target-consensus/AAPL": _load("fmp_price_target_aapl.json"),
    }


class FakeFmp(FmpClient):
    """FmpClient con `_fetch` inyectado desde fixtures (sin red real)."""

    def __init__(self, bundle=None, empty=()):
        self.api_key = "fake-key"  # is_available True, pero no hay red
        self.base_url = self.BASE_URL
        self.bundle = bundle if bundle is not None else _fmp_responses()
        self.empty = set(empty)  # endpoints que "no devuelven datos"
        self.calls = []

    def is_available(self):
        return True

    def _fetch(self, endpoint, params=None):
        self.calls.append(endpoint)
        if endpoint in self.empty:
            return []  # FMP devuelve lista vacía para "sin datos"
        return self.bundle.get(endpoint)


class SimpleFin:
    """Fake de Finnhub para el cruce (get_fundamentals disponible o no)."""

    def __init__(self, available=True, data=None):
        self._avail = available
        self._data = data or {"pe_ratio": 28.5, "roeTTM": 0.35}

    def is_available(self):
        return self._avail

    def get_fundamentals(self, symbol):
        return dict(self._data)


# ----------------------------------------------------------- FmpClient


def test_fmp_not_available_without_key(monkeypatch):
    # Aislar del .env real: en el repo principal puede existir una FMP_API_KEY
    # cargada (post-merge 30/08), y este test verifica la rama "sin key"
    # (FmpClient sin api_key -> is_available False -> _fetch None).
    monkeypatch.delenv("FMP_API_KEY", raising=False)
    monkeypatch.setattr("app.core.fundamentals_ingestion.settings.FMP_API_KEY", "")
    c = FmpClient()
    assert c.is_available() is False
    assert c._fetch("income-statement/AAPL") is None


def test_fmp_endpoints_pass_correct_path():
    c = FmpClient(api_key="fake")
    c._fetch = lambda ep, p=None: ep  # sin tocar red
    assert c.income_statement("AAPL") == "income-statement/AAPL"
    assert c.balance_sheet("AAPL") == "balance-sheet-statement/AAPL"
    assert c.cash_flow("AAPL") == "cash-flow-statement/AAPL"
    assert c.profile("AAPL") == "profile/AAPL"
    assert c.price_target_consensus("AAPL") == "price-target-consensus/AAPL"


def _make(tmp_path):
    """Instancia FundamentalsIngestion con FakeFmp + cache en tmp_path."""
    fake = FakeFmp()
    return FundamentalsIngestion(fmp=fake, cache_dir=str(tmp_path)), fake


def test_ingest_cache_miss_full_ingest(tmp_path, capsys):
    ing, _ = _make(tmp_path)
    payload = ing.ingest_symbol("AAPL")
    assert payload is not None
    assert payload["symbol"] == "AAPL"
    assert payload["_data_source"] == "fmp_live"
    assert len(payload["income_statement"]) == 2
    assert isinstance(payload["profile"], dict)
    assert isinstance(payload["price_target_consensus"], dict)
    assert "cache miss: full ingest" in capsys.readouterr().out
    cache_file = os.path.join(str(tmp_path), "AAPL.json")
    assert os.path.exists(cache_file)
    reloaded = json.load(open(cache_file))
    assert reloaded["symbol"] == "AAPL"
    assert len(reloaded["income_statement"]) == 2


def test_cache_hit_no_refresh(tmp_path, capsys):
    ing, fake = _make(tmp_path)
    ing.ingest_symbol("AAPL")  # miss inicial
    capsys.readouterr().out
    calls_after_miss = list(fake.calls)

    ing.ingest_symbol("AAPL")  # segundo: cache hit, sin re-llamar FMP
    out2 = capsys.readouterr().out
    assert "cache hit: fresh" in out2
    assert "no refresh needed" in out2
    assert fake.calls == calls_after_miss  # no hubo nuevas llamadas


def test_stale_cache_triggers_refresh(tmp_path, capsys):
    ing, _ = _make(tmp_path)
    ing.ingest_symbol("AAPL")
    capsys.readouterr().out
    path = os.path.join(str(tmp_path), "AAPL.json")
    old = time.time() - (TTL_DAYS + 1) * 86400
    os.utime(path, (old, old))
    assert ing.needs_refresh("AAPL") is True

    ing.ingest_symbol("AAPL")
    out = capsys.readouterr().out
    assert "cache stale" in out and "refreshing" in out
    assert "refresh: wrote" in out
    # se reescribió el cache (mtime reciente)
    assert time.time() - os.path.getmtime(path) < 300


def test_refresh_empty_preserves_stale(tmp_path, capsys):
    ing, fake = _make(tmp_path)
    ing.ingest_symbol("AAPL")
    capsys.readouterr().out
    path = os.path.join(str(tmp_path), "AAPL.json")
    old = time.time() - (TTL_DAYS + 1) * 86400
    os.utime(path, (old, old))
    # simular que en este refresh FMP no devuelve statements
    fake.empty = {"income-statement/AAPL", "cash-flow-statement/AAPL"}

    payload = ing.ingest_symbol("AAPL")
    # conserva el cache previo marcado stale, nunca lo borra
    assert payload is not None
    assert payload["_data_source"] == "stale_cache"
    assert os.path.exists(path)


def test_force_ignores_fresh_cache(tmp_path, capsys):
    ing, _ = _make(tmp_path)
    ing.ingest_symbol("AAPL")
    capsys.readouterr().out
    payload = ing.ingest_symbol("AAPL", force=True)
    assert payload is not None
    assert payload["_data_source"] == "fmp_live"


def test_empty_corrupt_cache_treated_as_miss(tmp_path, capsys):
    ing, _ = _make(tmp_path)
    path = os.path.join(str(tmp_path), "AAPL.json")
    with open(path, "w") as f:
        f.write("{esto-no-es-json")  # corrupto
    payload = ing.ingest_symbol("AAPL")
    assert payload is not None
    assert payload["_data_source"] == "fmp_live"
    assert "cache empty/corrupt" in capsys.readouterr().out


def test_finnhub_cross_unverified_flagged(tmp_path):
    ing, _ = _make(tmp_path)
    ing.finnhub = SimpleFin(available=True, data={"pe_ratio": 28.5})
    payload = ing.ingest_symbol("AAPL")
    assert payload["finnhub_crosscheck"] is not None
    assert payload["finnhub_crosscheck"]["_cross_unverified"] is True


def test_finnhub_cross_skipped_when_unavailable(tmp_path):
    ing, _ = _make(tmp_path)
    ing.finnhub = SimpleFin(available=False)
    payload = ing.ingest_symbol("AAPL")
    assert payload["finnhub_crosscheck"] is None


def test_needs_refresh_false_when_fresh(tmp_path):
    ing, _ = _make(tmp_path)
    ing.ingest_symbol("AAPL")
    assert ing.needs_refresh("AAPL") is False


# ---------------- integración: fixture Excel real (formato/columnas reales) ----

# Mirror del motor canónico (motor_screening.py). La Fase 2/3 debe producir
# EXACTAMENTE esta surface a partir de los datos crudos de Fase 1. Mantener
# sincronizado con ~/.claude/skills/aai-screening-acciones/scripts/motor_screening.py
COLS_ESENCIALES = ["Name", "Ticker", "Price, Current"]
COLS_NUCLEO = [
    "Fair Value", "Fair Value Label (Analyst Targets)", "Overall Health Label",
    "EV / EBIT", "Free Cash Flow Yield",
    "Return on Invested Capital", "Avg Return on Invested Capital (5y)",
    "Return on Equity", "Gross Profit Margin",
    "Avg EPS Growth (5y)", "Revenue CAGR (5y)",
    "FCF / Net Income", "Buyback Yield",
    "Piotroski Score", "Altman Z-Score", "Beneish M-Score",
    "Total Debt / Total Capital",
]

# Ruta por defecto al export real del screener de InvestingPro (reference).
# Vive DENTRO del repo (no en ~/Downloads, directorio volátil donde se perdió
# y dejó este test en skip silencioso). Path relativo al archivo de test.
# Sobrescribible por env REAL_EXCEL_FIXTURE para otros equipos/CI.
_CANON_FIXTURE_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "fixtures", "canon"
)
_DEFAULT_CANON_XLSX = os.path.join(_CANON_FIXTURE_DIR, "market_view_export.xlsx")
REAL_EXCEL = os.environ.get("REAL_EXCEL_FIXTURE", _DEFAULT_CANON_XLSX)


def _load_excel_headers(path):
    """Devuelve (fila_encabezados, [nombres]). Encuentra la fila que tiene
    'Name' y 'Ticker' (regla dura del motor de AAI, orden libre)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers = []
    for row in ws.iter_rows(values_only=True):
        vals = [(" ".join(str(v).split()) if v is not None else "") for v in row]
        if "Name" in vals and "Ticker" in vals:
            header_row = vals
            break
    wb.close()
    if header_row is None:
        raise ValueError("no se encontró fila de encabezados (debe contener Name y Ticker)")
    return [h for h in header_row if h]


def test_excel_fixture_covers_motor_core_columns():
    """El export real cubre las columnas núcleo y esenciales del motor canónico.
    Garantiza que la surface de datos de Fase 1+2 puede rellenar TODAS las
    columnas que el motor espera (hmap), sin re-diseñar nada.

    RUIDOSO: si el fixture falta, skip con advertencia visible (nunca
    silencioso). Con REQUIRE_PARIDAD=1, falla en rojo."""
    if not os.path.exists(REAL_EXCEL):
        msg = (
            f"PARIDAD NO VERIFICADA: falta el export real en {REAL_EXCEL}. "
            f"Re-exportá desde InvestingPro y copialo a "
            f"{_CANON_FIXTURE_DIR}/market_view_export.xlsx."
        )
        if os.environ.get("REQUIRE_PARIDAD") == "1":
            pytest.fail(msg)
        import warnings
        warnings.warn(f"⚠️  {msg}", stacklevel=2)
        pytest.skip(msg)
    headers = _load_excel_headers(REAL_EXCEL)
    for col in COLS_ESENCIALES + COLS_NUCLEO:
        assert col in headers, f"columna núcleo ausente en export real: {col}"